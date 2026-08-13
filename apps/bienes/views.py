from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q, Count
from django.utils import timezone
from django.http import JsonResponse
from openpyxl import load_workbook
import xlrd

from .models import Bien
from apps.core.models import Sede, Area
from apps.usuarios.models import Usuario
from .forms import BienForm
from .services.importacion_service import importar_desde_excel, generar_codigo_patrimonial
from apps.movimientos.models import MovimientoBien
from django.db.models import Prefetch


# ============================
# LISTADO DE BIENES
# ============================
@login_required
def lista_bienes(request):
    user = request.user
    query = request.GET.get('q', '').strip()
    area_id = request.GET.get('area', '').strip()

    es_admin = user.es_inventariador_o_admin

    if es_admin:
        # 1. ADMIN: Cargar sedes y pre-contar bienes por área
        areas_prefetch = Area.objects.annotate(
            total_bienes_area=Count('bien', filter=Q(bien__activo=True))
        ).select_related('sede')

        sedes = Sede.objects.prefetch_related(
            Prefetch('areas', queryset=areas_prefetch)
        ).all()

        bienes_qs = Bien.objects.filter(activo=True).select_related('area', 'area__sede', 'usuario_responsable')
        total_bienes_sistema = bienes_qs.count()
        areas_usuario = None
    else:
        sedes = None
        total_bienes_sistema = None

        # 1. CUSTODIO: Bienes bajo su responsabilidad
        bienes_qs = Bien.objects.filter(
            usuario_responsable=user,
            activo=True
        ).select_related('area', 'area__sede')

        # 2. Áreas específicas donde el usuario tiene bienes
        areas_usuario = Area.objects.filter(
            bien__usuario_responsable=user,
            bien__activo=True
        ).annotate(
            total_bienes_area=Count('bien', filter=Q(bien__usuario_responsable=user, bien__activo=True))
        ).select_related('sede').distinct()

    # Filtros
    if area_id:
        bienes_qs = bienes_qs.filter(area_id=area_id)

    if query:
        bienes_qs = bienes_qs.filter(
            Q(codigo_patrimonial__icontains=query) |
            Q(denominacion__icontains=query) |
            Q(marca__icontains=query) |
            Q(serie__icontains=query) |
            Q(catalogo_siga__icontains=query)
        )

    paginator = Paginator(bienes_qs.order_by('-creado_en'), 15)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # AQUÍ USAMOS EL RELATED_NAME CORRECTO: 'bienes_actuales'
    usuarios_activos = Usuario.objects.filter(is_active=True).prefetch_related(
        Prefetch(
            'bienes_actuales',
            queryset=Bien.objects.filter(activo=True).select_related('area', 'area__sede')
        )
    ).order_by('last_name')

    context = {
        'es_admin': es_admin,
        'page_obj': page_obj,
        'total_bienes_sistema': total_bienes_sistema,
        'total_bienes_custodia': bienes_qs.count() if not es_admin else None,
        'sedes': sedes,
        'areas_usuario': areas_usuario,
        'query': query,
        'area_seleccionada': area_id,
        'usuarios_activos': usuarios_activos,
        'areas': Area.objects.select_related('sede').all().order_by('nombre'),
    }

    return render(request, 'bienes/lista.html', context)

    return render(request, 'bienes/lista.html', context)


# ============================
# IMPRESIÓN MASIVA DE QR (HOJA A4)
# ============================
@login_required
def imprimir_qrs_area(request, area_id):
    area = get_object_or_404(Area, id=area_id)

    # Obtener todos los bienes con QR generado de esta área
    bienes = Bien.objects.filter(area=area, activo=True).exclude(qr_code='').order_by('codigo_patrimonial')

    context = {
        'area': area,
        'bienes': bienes,
    }
    return render(request, 'bienes/imprimir_qrs_a4.html', context)


# ============================
# REGISTRO DE BIEN
# ============================
@login_required
def registro_bien(request):
    if not request.user.es_inventariador_o_admin:
        messages.error(request, " Acceso restringido. Solo el personal de Control Patrimonial puede registrar bienes.")
        return redirect('bienes:lista')

    if request.method == 'POST':
        form = BienForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                bien = form.save(commit=False)

                area = bien.area
                bien.codigo_patrimonial = generar_codigo_patrimonial(area.nombre if area else "GENERAL")

                bien.creado_por = request.user
                bien.full_clean()
                bien.save()

                bien.generar_qr()
                bien.save(update_fields=['qr_code'])

                MovimientoBien.objects.create(
                    bien=bien,
                    tipo='ASIGNACION',
                    area_destino=bien.area,
                    usuario_destino=bien.usuario_responsable,
                    fecha_movimiento=timezone.now(),
                    registrado_por=request.user
                )

                messages.success(request, f" Bien {bien.codigo_patrimonial} registrado exitosamente.")
                return redirect('bienes:detalle', codigo=bien.codigo_patrimonial)
        else:
            messages.error(request, " Revise los campos del formulario.")
    else:
        form = BienForm()

    return render(request, 'bienes/registro.html', {'form': form})


# ============================
# DETALLE PÚBLICO (DESDE QR)
# ============================
def detalle_publico(request, codigo):
    bien = get_object_or_404(Bien, codigo_patrimonial=codigo, activo=True)
    return render(request, 'bienes/detalle_publico.html', {'bien': bien})


# ============================
# DETALLE CON AUTENTICACIÓN
# ============================
@login_required
def detalle_bien(request, codigo):
    bien = get_object_or_404(Bien, codigo_patrimonial=codigo)
    return render(request, 'bienes/detalle.html', {'bien': bien})


# ============================
# EDITAR BIEN
# ============================
@login_required
def editar_bien(request, codigo):
    if not request.user.es_inventariador_o_admin:
        messages.error(request, " No tiene permisos para modificar la información patrimonial.")
        return redirect('bienes:lista')

    bien = get_object_or_404(Bien, codigo_patrimonial=codigo)
    if request.method == 'POST':
        form = BienForm(request.POST, instance=bien)
        if form.is_valid():
            bien = form.save()
            messages.success(request, f" Bien {bien.codigo_patrimonial} actualizado correctamente.")
            return redirect('bienes:detalle', codigo=bien.codigo_patrimonial)
    else:
        form = BienForm(instance=bien)
    return render(request, 'bienes/registro.html', {'form': form, 'edicion': True})


# ============================
# DAR DE BAJA UN BIEN
# ============================
@login_required
def baja_bien(request, codigo):
    if not request.user.es_inventariador_o_admin:
        messages.error(request, " No tiene permisos para procesar bajas de bienes.")
        return redirect('bienes:lista')

    bien = get_object_or_404(Bien, codigo_patrimonial=codigo)
    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()
        if not motivo:
            messages.error(request, " Debe especificar un motivo de baja.")
            return redirect('bienes:detalle', codigo=bien.codigo_patrimonial)

        with transaction.atomic():
            bien.activo = False
            bien.fecha_baja = timezone.now().date()
            bien.motivo_baja = motivo
            bien.save()

            MovimientoBien.objects.create(
                bien=bien,
                tipo='BAJA',
                area_origen=bien.area,
                usuario_origen=bien.usuario_responsable,
                fecha_movimiento=timezone.now(),
                observaciones=motivo,
                registrado_por=request.user
            )

        messages.warning(request, f"⚠ Bien {bien.codigo_patrimonial} dado de baja exitosamente.")
        return redirect('bienes:lista')
    return render(request, 'bienes/confirmar_baja.html', {'bien': bien})


# ============================
# IMPORTACIÓN MASIVA EXCEL
# ============================
@login_required
def importar_bienes(request):
    if not request.user.es_inventariador_o_admin:
        messages.error(request, " No tiene permisos para realizar importaciones masivas.")
        return redirect('bienes:lista')

    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        hojas_seleccionadas = request.POST.getlist('hojas')

        if not hojas_seleccionadas:
            messages.error(request, " Debe seleccionar al menos una hoja para importar.")
            return redirect('bienes:importar')

        nombre_archivo = archivo.name.lower()
        es_xlsx = nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xlsm')

        if es_xlsx:
            wb = load_workbook(archivo, data_only=True)
            hojas_disponibles = wb.sheetnames
        else:
            wb = xlrd.open_workbook(file_contents=archivo.read(), formatting_info=False)
            hojas_disponibles = wb.sheet_names()

        hojas_validas = [h for h in hojas_seleccionadas if h in hojas_disponibles]
        if not hojas_validas:
            messages.error(request, " Ninguna de las hojas seleccionadas es válida.")
            return redirect('bienes:importar')

        try:
            resultado = importar_desde_excel(archivo, request.user, hojas_validas)
            messages.success(request, f" Importación completada: {resultado['bienes_creados']} bienes creados.")
            if resultado['errores']:
                for error in resultado['errores'][:10]:
                    messages.warning(request, f"️ {error}")
        except Exception as e:
            messages.error(request, f" Error en la importación: {str(e)}")
        return redirect('bienes:lista')

    return render(request, 'importacion/importar.html')


@login_required
def listar_hojas(request):
    if request.method == 'POST' and request.FILES.get('archivo_excel'):
        archivo = request.FILES['archivo_excel']
        nombre_archivo = archivo.name.lower()
        es_xlsx = nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xlsm')

        try:
            if es_xlsx:
                wb = load_workbook(archivo, data_only=True)
                hojas = wb.sheetnames
            else:
                wb = xlrd.open_workbook(file_contents=archivo.read(), formatting_info=False)
                hojas = wb.sheet_names()

            return JsonResponse({'hojas': hojas, 'importadas': []})
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=400)

    return JsonResponse({'error': 'Método no permitido'}, status=405)


# ============================
# CONSULTA DE BIENES EN CUSTODIA
# ============================
@login_required
def mis_bienes(request):
    user = request.user
    mis_areas = user.areas_custodia.all()

    bienes = Bien.objects.filter(
        Q(usuario_responsable=user) | Q(area__in=mis_areas),
        activo=True
    ).distinct().select_related('area', 'usuario_responsable')

    return render(request, 'bienes/mis_bienes.html', {
        'bienes': bienes,
        'mis_areas': mis_areas
    })

@login_required
def asignacion_masiva(request):
    # Verificación de permisos
    if not getattr(request.user, 'es_inventariador_o_admin', False) and not request.user.is_staff:
        messages.error(request, " No tiene permisos para realizar asignaciones masivas.")
        return redirect('bienes:lista')

    if request.method == 'POST':
        bienes_ids_str = request.POST.get('bienes_ids', '')
        usuario_destino_id = request.POST.get('usuario_destino')
        area_destino_id = request.POST.get('area_destino')

        if not bienes_ids_str or not usuario_destino_id or not area_destino_id:
            messages.error(request, " Debe seleccionar bienes, usuario y área de destino.")
            return redirect('bienes:lista')

        # Convertir IDs a lista de enteros de forma segura
        try:
            bienes_ids = [int(id_str.strip()) for id_str in bienes_ids_str.split(',') if id_str.strip()]
        except ValueError:
            messages.error(request, " Formato de identificadores de bienes inválido.")
            return redirect('bienes:lista')

        # Validar existencia de destino
        try:
            usuario = Usuario.objects.get(id=usuario_destino_id)
            area = Area.objects.get(id=area_destino_id)
        except (Usuario.DoesNotExist, Area.DoesNotExist):
            messages.error(request, " El usuario o área seleccionada ya no existe.")
            return redirect('bienes:lista')

        bienes = Bien.objects.filter(id__in=bienes_ids, activo=True)

        if not bienes.exists():
            messages.warning(request, " No se encontraron bienes activos para reasignar.")
            return redirect('bienes:lista')

        try:
            with transaction.atomic():
                movimientos = []
                now = timezone.now()

                for bien in bienes:
                    # Crear el registro histórico de movimiento
                    movimientos.append(
                        MovimientoBien(
                            bien=bien,
                            tipo='TRANSFERENCIA',
                            area_origen=bien.area,
                            usuario_origen=bien.usuario_responsable,
                            area_destino=area,
                            usuario_destino=usuario,
                            fecha_movimiento=now,
                            documento_autorizacion='ASIGNACION_MASIVA_WEB',
                            observaciones=f"Asignación masiva realizada por {request.user.get_full_name() or request.user.username}.",
                            registrado_por=request.user
                        )
                    )
                    # Actualizar las propiedades del bien
                    bien.usuario_responsable = usuario
                    bien.area = area
                    bien.sede = area.sede

                # Guardar movimientos en bloque
                MovimientoBien.objects.bulk_create(movimientos)

                # Actualizar bienes en bloque
                Bien.objects.bulk_update(bienes, fields=['usuario_responsable', 'area', 'sede'])

            messages.success(
                request,
                f" Se reasignaron con éxito {len(bienes)} bienes a {usuario.get_full_name()} en el área {area.nombre}."
            )
        except Exception as e:
            messages.error(request, f" Ocurrió un error al procesar la transacción: {str(e)}")

    return redirect('bienes:lista')


@login_required
def buscar_bienes_ajax(request):
    """Endpoint ligero para autocompletado de bienes."""
    query = request.GET.get('q', '').strip()

    if len(query) < 2:
        return JsonResponse({'results': []})

    bienes = Bien.objects.filter(
        Q(codigo_patrimonial__icontains=query) |
        Q(denominacion__icontains=query) |
        Q(serie__icontains=query),
        activo=True
    ).select_related('area', 'sede')[:20]  # Limitamos a los primeros 20 resultados

    results = [
        {
            'id': bien.id,
            'text': f"{bien.codigo_patrimonial} - {bien.denominacion}",
            'area': f"{bien.sede.nombre} / {bien.area.nombre}"
        }
        for bien in bienes
    ]

    return JsonResponse({'results': results})