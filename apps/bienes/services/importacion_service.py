# bienes/services/importacion_service.py
import re
from decimal import Decimal
from datetime import datetime
from django.db import transaction
from django.core.exceptions import ValidationError
from openpyxl import load_workbook
import xlrd

from apps.core.models import Sede, Area, Cargo
from apps.usuarios.models import Usuario
from apps.bienes.models import Bien
from apps.movimientos.models import MovimientoBien


def valor_limpio(valor):
    """Convierte un valor a string y elimina espacios, o retorna cadena vacía si es None."""
    if valor is None:
        return ''
    return str(valor).strip()

# ===== FUNCIÓN PRINCIPAL =====
def importar_desde_excel(archivo_excel, usuario_actual, hojas_seleccionadas=None):
    """
    Importa bienes desde un archivo Excel (.xls o .xlsx).
    - hojas_seleccionadas: lista de nombres de hojas a procesar (si es None, procesa todas).
    - Cada bien se guarda en su propia transacción.
    - Las áreas deben existir previamente (creadas manualmente).
    - El código patrimonial se genera automáticamente con formato BN-AREA-AÑO-SECUENCIA.
    - El código SIGA del Excel se guarda en el campo catalogo_siga.
    - SATISFACE: QA-2 (Integridad) - Si un bien falla, no afecta a los demás.
    """
    nombre_archivo = archivo_excel.name.lower()
    es_xlsx = nombre_archivo.endswith('.xlsx') or nombre_archivo.endswith('.xlsm')

    # Cargar el libro
    if es_xlsx:
        wb = load_workbook(archivo_excel, data_only=True)
        hojas = wb.sheetnames
    else:
        wb = xlrd.open_workbook(file_contents=archivo_excel.read(), formatting_info=False)
        hojas = wb.sheet_names()

    resultados = {
        'total_hojas': 0,
        'bienes_creados': 0,
        'usuarios_encontrados': 0,
        'errores': []
    }

    # === 1. OBTENER LA SEDE EXISTENTE (creada manualmente) ===
    sede = Sede.objects.filter(nombre="IESTP ARIB").first()
    if not sede:
        raise ValueError("La sede 'IESTP ARIB' no existe. Cree la sede manualmente antes de importar.")

    # === 2. USUARIO GENÉRICO (para responsables no existentes) ===
    usuario_generico = Usuario.objects.filter(dni='00000000').first()
    if not usuario_generico:
        cargo, _ = Cargo.objects.get_or_create(nombre="SIN CARGO")
        area_generica, _ = Area.objects.get_or_create(
            nombre="ALMACEN CENTRAL",
            defaults={'sede': sede}
        )
        usuario_generico = Usuario.objects.create(
            username='sin_asignar',
            dni='00000000',
            first_name='SIN',
            last_name='ASIGNAR',
            email='sin_asignar@iestparib.edu.pe',
            cargo=cargo,
            area=area_generica,
            is_active=True
        )

    # === FILTRAR HOJAS SEGÚN SELECCIÓN ===
    if hojas_seleccionadas is None:
        hojas_a_procesar = hojas
    else:
        hojas_a_procesar = [h for h in hojas if h in hojas_seleccionadas]

    # === 3. PROCESAR CADA HOJA SELECCIONADA ===
    for sheet_name in hojas_a_procesar:
        resultados['total_hojas'] += 1

        # Extraer datos de la hoja
        if es_xlsx:
            sheet = wb[sheet_name]
            data = extraer_datos_xlsx(sheet)
        else:
            sheet = wb.sheet_by_name(sheet_name)
            data = extraer_datos_xls(sheet)

        responsable_data = data['responsable']
        if not responsable_data or not responsable_data.get('dni'):
            resultados['errores'].append(f"Hoja '{sheet_name}': No se encontró DNI del responsable.")
            continue

        dni = responsable_data['dni']
        # Buscar usuario por DNI
        usuario = Usuario.objects.filter(dni=dni).first()
        if not usuario:
            usuario = usuario_generico
            resultados['errores'].append(
                f"Hoja '{sheet_name}': Usuario con DNI {dni} no existe. Asignado a 'SIN ASIGNAR'."
            )
        else:
            resultados['usuarios_encontrados'] += 1

        # === 4. BUSCAR ÁREA EXISTENTE (NO CREAR) ===
        area_nombre = sheet_name.strip()
        area = Area.objects.filter(nombre=area_nombre).first()
        if not area:
            resultados['errores'].append(
                f"Hoja '{sheet_name}': El área '{area_nombre}' no existe en el sistema. "
                f"Cree el área manualmente antes de importar."
            )
            continue  # Saltar esta hoja

        # === 5. PROCESAR FILAS DE BIENES ===
        filas = data['filas']
        for row in filas:
            if not row or not row.get('denominacion'):
                continue

            cantidad = row.get('cantidad', 1)
            for i in range(cantidad):
                with transaction.atomic():
                    try:
                        # Crear el bien
                        bien = Bien(
                            denominacion=row.get('denominacion', ''),
                            marca=row.get('marca', ''),
                            modelo=row.get('modelo', ''),
                            serie=row.get('serie', ''),
                            color=row.get('color', ''),
                            dimensiones=row.get('dimensiones', ''),
                            estado_conservacion=row.get('estado_conservacion', 'B'),
                            observaciones=row.get('observaciones', ''),
                            sede=sede,
                            area=area,
                            usuario_responsable=usuario,
                            tipo_adquisicion='',
                            numero_documento=f"IMP-{sheet_name[:10]}",
                            fecha_adquisicion=datetime.now().date(),
                            valor_inicial=Decimal('0.00'),
                            depreciable=False,
                        )

                        # === GENERAR CÓDIGO PATRIMONIAL CON FORMATO BN-AREA-AÑO-SECUENCIA ===
                        bien.codigo_patrimonial = generar_codigo_patrimonial(area_nombre)

                        # === GUARDAR CÓDIGO SIGA DEL EXCEL (si existe) ===
                        codigo_siga = row.get('codigo_patrimonial')
                        if codigo_siga is not None and str(codigo_siga).strip():
                            bien.catalogo_siga = str(codigo_siga).strip()

                        # === codigo_interno se genera automáticamente (UUIDField) ===
                        bien.creado_por = usuario_actual
                        bien.full_clean()
                        bien.save()

                        # === GENERAR QR ===
                        try:
                            bien.generar_qr()
                            bien.save(update_fields=['qr_code'])
                        except Exception as qr_error:
                            resultados['errores'].append(
                                f"Hoja '{sheet_name}': Error al generar QR para {bien.codigo_patrimonial}: {str(qr_error)}"
                            )

                        resultados['bienes_creados'] += 1

                        # === REGISTRAR MOVIMIENTO DE ASIGNACIÓN ===
                        MovimientoBien.objects.create(
                            bien=bien,
                            tipo='ASIGNACION',
                            area_destino=area,
                            usuario_destino=usuario,
                            fecha_movimiento=datetime.now(),
                            registrado_por=usuario_actual
                        )
                    except Exception as e:
                        resultados['errores'].append(
                            f"Hoja '{sheet_name}': Error al crear bien: {str(e)}"
                        )

    return resultados


# ===== FUNCIONES AUXILIARES DE EXTRACCIÓN =====

def extraer_datos_xlsx(sheet):
    """
    Extrae responsable y filas de bienes con búsqueda robusta.
    SATISFACE: QA-6 (Interoperabilidad) - Tolerante a diferentes formatos.
    """
    responsable = {}
    filas = []

    # 1. Buscar la fila de cabecera (iterar sobre todas las filas)
    header_row = None
    header_idx = None
    for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
        if not row:
            continue
        texto = ' '.join(str(cell).upper() for cell in row if cell)
        # Palabras clave para identificar la cabecera
        if 'ITEM' in texto and any(k in texto for k in ['CODIGO', 'DENOMINACION', 'DESCRIPCION', 'MARCA']):
            header_row = row
            header_idx = i
            break

    # Si no se encontró, buscar una fila con "DESCRIPCION DEL BIEN" y tomar la siguiente
    if not header_row:
        for i, row in enumerate(sheet.iter_rows(min_row=1, values_only=True), start=1):
            if not row:
                continue
            texto = ' '.join(str(cell).upper() for cell in row if cell)
            if 'DESCRIPCION DEL BIEN' in texto:
                next_row = list(sheet.iter_rows(min_row=i+1, max_row=i+1, values_only=True))[0]
                if next_row:
                    header_row = next_row
                    header_idx = i + 1
                break

    if not header_row:
        return {'responsable': responsable, 'filas': []}

    # 2. Mapear nombres de columna a índices (normalizar nombres)
    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell:
            name = str(cell).strip().upper().replace('.', '').replace(' ', '_')
            col_map[name] = idx

    # 3. Buscar responsable en todas las filas ANTES de la cabecera
    for i in range(1, header_idx):
        row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            cell_str = str(cell).strip()
            # Buscar "NOMBRES Y APELLIDOS" o "NOMBRE Y APELLIDOS"
            if re.search(r'NOMBRES?\s+Y\s+APELLIDOS', cell_str, re.IGNORECASE):
                if ':' in cell_str:
                    partes = cell_str.split(':', 1)
                    if len(partes) > 1:
                        responsable['nombre'] = partes[1].strip()
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    responsable['nombre'] = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
            elif re.search(r'D\.?N\.?I\.?', cell_str, re.IGNORECASE):
                dni_match = re.search(r'D\.?N\.?I\.?\s*[:]\s*(\d{8})', cell_str, re.IGNORECASE)
                if dni_match:
                    responsable['dni'] = dni_match.group(1)
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    dni_val = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
                    dni_match = re.search(r'\d{8}', dni_val)
                    if dni_match:
                        responsable['dni'] = dni_match.group()
            elif re.search(r'CARGO', cell_str, re.IGNORECASE):
                if ':' in cell_str:
                    partes = cell_str.split(':', 1)
                    if len(partes) > 1:
                        responsable['cargo'] = partes[1].strip()
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    responsable['cargo'] = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
            if responsable.get('nombre') and responsable.get('dni'):
                break
        if responsable.get('nombre') and responsable.get('dni'):
            break

    # Si no se encontró DNI, intentar buscar en toda la hoja (por si el formato es diferente)
    if not responsable.get('dni'):
        for i in range(1, header_idx):
            row = list(sheet.iter_rows(min_row=i, max_row=i, values_only=True))[0]
            if not row:
                continue
            row_text = ' '.join(str(cell) for cell in row if cell)
            dni_match = re.search(r'\b(\d{8})\b', row_text)
            if dni_match:
                responsable['dni'] = dni_match.group(1)
                break

    # 4. Leer filas de bienes (después de la cabecera)
    for row in sheet.iter_rows(min_row=header_idx + 1, values_only=True):
        if not row or not any(row):
            continue
        texto_fila = ' '.join(str(cell).upper() for cell in row if cell)
        if 'ITEM' in texto_fila or 'CODIGO' in texto_fila or 'DESCRIPCION DEL BIEN' in texto_fila:
            continue

        # Obtener denominación (usar el mapa de columnas o índice por defecto)
        denominacion = row[col_map.get('DENOMINACION', 3)] if col_map.get('DENOMINACION', 3) < len(row) else ''
        if not denominacion:
            denominacion = row[col_map.get('DESCRIPCION', 3)] if col_map.get('DESCRIPCION', 3) < len(row) else ''
        if not denominacion:
            continue

        cantidad_raw = row[col_map.get('CANT.', 6)] if col_map.get('CANT.', 6) < len(row) else 1
        try:
            cantidad = int(cantidad_raw) if cantidad_raw else 1
        except (ValueError, TypeError):
            cantidad = 1

        # Función auxiliar para obtener un valor limpio (None -> cadena vacía)
        def obtener_valor(indice):
            if indice < len(row):
                return str(row[indice]).strip() if row[indice] is not None else ''
            return ''

        bien_data = {
            'denominacion': denominacion,
            'codigo_patrimonial': obtener_valor(col_map.get('CODIGO_PATRIMONIAL', 2)),
            'marca': obtener_valor(col_map.get('MARCA', 4)),
            'modelo': obtener_valor(col_map.get('MODELO', 5)),
            'cantidad': cantidad,
            'color': obtener_valor(col_map.get('COLOR', 7)),
            'dimensiones': obtener_valor(col_map.get('DIMENSION', 8)),
            'serie': obtener_valor(col_map.get('SERIE', 9)),
            'estado_conservacion': mapear_estado(obtener_valor(col_map.get('ESTADO_DE_CONSERV', 10))),
            'observaciones': obtener_valor(col_map.get('OBSERVACIONES', 11)),
        }
        filas.append(bien_data)

    return {'responsable': responsable, 'filas': filas}


def extraer_datos_xls(sheet):
    """
    Versión para .xls usando xlrd.
    """
    responsable = {}
    filas = []

    # Buscar cabecera
    header_row = None
    header_idx = None
    for row_idx in range(sheet.nrows):
        row = sheet.row_values(row_idx)
        if not row:
            continue
        texto = ' '.join(str(cell).upper() for cell in row if cell)
        if 'ITEM' in texto and any(k in texto for k in ['CODIGO', 'DENOMINACION', 'DESCRIPCION', 'MARCA']):
            header_row = row
            header_idx = row_idx
            break

    if not header_row:
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            if not row:
                continue
            texto = ' '.join(str(cell).upper() for cell in row if cell)
            if 'DESCRIPCION DEL BIEN' in texto:
                if row_idx + 1 < sheet.nrows:
                    header_row = sheet.row_values(row_idx + 1)
                    header_idx = row_idx + 1
                break

    if not header_row:
        return {'responsable': responsable, 'filas': []}

    col_map = {}
    for idx, cell in enumerate(header_row):
        if cell:
            name = str(cell).strip().upper().replace('.', '').replace(' ', '_')
            col_map[name] = idx

    # Buscar responsable
    for row_idx in range(header_idx):
        row = sheet.row_values(row_idx)
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            if not cell:
                continue
            cell_str = str(cell).strip()
            if re.search(r'NOMBRES?\s+Y\s+APELLIDOS', cell_str, re.IGNORECASE):
                if ':' in cell_str:
                    partes = cell_str.split(':', 1)
                    if len(partes) > 1:
                        responsable['nombre'] = partes[1].strip()
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    responsable['nombre'] = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
            elif re.search(r'D\.?N\.?I\.?', cell_str, re.IGNORECASE):
                dni_match = re.search(r'D\.?N\.?I\.?\s*[:]\s*(\d{8})', cell_str, re.IGNORECASE)
                if dni_match:
                    responsable['dni'] = dni_match.group(1)
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    dni_val = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
                    dni_match = re.search(r'\d{8}', dni_val)
                    if dni_match:
                        responsable['dni'] = dni_match.group()
            elif re.search(r'CARGO', cell_str, re.IGNORECASE):
                if ':' in cell_str:
                    partes = cell_str.split(':', 1)
                    if len(partes) > 1:
                        responsable['cargo'] = partes[1].strip()
                elif col_idx + 1 < len(row) and row[col_idx + 1]:
                    responsable['cargo'] = str(row[col_idx + 1]).strip() if row[col_idx + 1] else ''
            if responsable.get('nombre') and responsable.get('dni'):
                break
        if responsable.get('nombre') and responsable.get('dni'):
            break

    if not responsable.get('dni'):
        for row_idx in range(header_idx):
            row = sheet.row_values(row_idx)
            row_text = ' '.join(str(cell) for cell in row if cell)
            dni_match = re.search(r'\b(\d{8})\b', row_text)
            if dni_match:
                responsable['dni'] = dni_match.group(1)
                break

    # Leer filas de datos
    for row_idx in range(header_idx + 1, sheet.nrows):
        row = sheet.row_values(row_idx)
        if not any(row):
            continue
        texto_fila = ' '.join(str(cell).upper() for cell in row if cell)
        if 'ITEM' in texto_fila or 'CODIGO' in texto_fila or 'DESCRIPCION DEL BIEN' in texto_fila:
            continue

        denominacion = row[col_map.get('DENOMINACION', 3)] if col_map.get('DENOMINACION', 3) < len(row) else ''
        if not denominacion:
            denominacion = row[col_map.get('DESCRIPCION', 3)] if col_map.get('DESCRIPCION', 3) < len(row) else ''
        if not denominacion:
            continue

        cantidad_raw = row[col_map.get('CANT.', 6)] if col_map.get('CANT.', 6) < len(row) else 1
        try:
            cantidad = int(cantidad_raw) if cantidad_raw else 1
        except (ValueError, TypeError):
            cantidad = 1

        def obtener_valor(indice):
            if indice < len(row):
                return str(row[indice]).strip() if row[indice] is not None else ''
            return ''

        bien_data = {
            'denominacion': denominacion,
            'codigo_patrimonial': obtener_valor(col_map.get('CODIGO_PATRIMONIAL', 2)),
            'marca': obtener_valor(col_map.get('MARCA', 4)),
            'modelo': obtener_valor(col_map.get('MODELO', 5)),
            'cantidad': cantidad,
            'color': obtener_valor(col_map.get('COLOR', 7)),
            'dimensiones': obtener_valor(col_map.get('DIMENSION', 8)),
            'serie': obtener_valor(col_map.get('SERIE', 9)),
            'estado_conservacion': mapear_estado(obtener_valor(col_map.get('ESTADO_DE_CONSERV', 10))),
            'observaciones': obtener_valor(col_map.get('OBSERVACIONES', 11)),
        }
        filas.append(bien_data)

    return {'responsable': responsable, 'filas': filas}


# ===== FUNCIONES AUXILIARES =====

def mapear_estado(valor):
    """Mapea texto de estado a código de una letra. Maneja None."""
    mapa = {
        'BUENO': 'B', 'REGULAR': 'R', 'MALO': 'M',
        'INSERVIBLE': 'I', 'NUEVO': 'N', 'MALOGRADO': 'M',
        'M.M': 'M', 'M.M.': 'M'
    }
    if not valor:
        return 'B'
    texto = str(valor).strip().upper()
    return mapa.get(texto, 'B')

def generar_codigo_patrimonial(area_nombre=None):
    """
    Genera código patrimonial único con formato BN-AREA-AÑO-SECUENCIA.
    Si no se proporciona área, usa "GENERAL".
    """
    from datetime import date
    import re
    año = date.today().year
    if area_nombre:
        area_clean = re.sub(r'[^A-Z0-9]', '', area_nombre.upper().strip())[:5]
    else:
        area_clean = "GENERAL"
    prefijo = f"BN-{area_clean}-{año}-"
    ultimo = Bien.objects.filter(codigo_patrimonial__startswith=prefijo).order_by('-codigo_patrimonial').first()
    if ultimo:
        try:
            secuencia = int(ultimo.codigo_patrimonial.split('-')[-1]) + 1
        except:
            secuencia = 1
    else:
        secuencia = 1
    return f"{prefijo}{secuencia:05d}"
#def generar_codigo_patrimonial(area_nombre):
#    """
#    Genera código patrimonial único con formato BN-AREA-AÑO-SECUENCIA.
#    Ejemplo: BN-BIBLIOTECA-2026-00001
#    """
#    from datetime import date
#    año = date.today().year
#    # Normalizar nombre del área (mayúsculas, sin espacios ni caracteres especiales)
#    area_clean = re.sub(r'[^A-Z0-9]', '', area_nombre.upper().strip())[:5]  # Truncar a 5 caracteres
#    prefijo = f"BN-{area_clean}-{año}-"
#    ultimo = Bien.objects.filter(
#        codigo_patrimonial__startswith=prefijo
#    ).order_by('-codigo_patrimonial').first()
#    if ultimo:
#        try:
#            secuencia = int(ultimo.codigo_patrimonial.split('-')[-1]) + 1
#        except:
#            secuencia = 1
#    else:
#        secuencia = 1
#    return f"{prefijo}{secuencia:05d}"